#a web application by sricharan
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template, request

app = Flask(__name__)

# Load the Excel file into a Pandas dataframe
df = pd.read_excel("raghuvamsha_sarga_03.xlsx")

# Define the columns we want to display on the web page
display_columns = ["Sloka_Number", "Sloka", "Sloka_Annotation", "Sloka_Translation", "Label"]

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get the Label value submitted by the user
        label_value = request.form.get("label")

        # Get the index of the row that was edited
        row_index = int(request.form.get("row_index"))

        # Update the corresponding cell in the Excel file
        wb = load_workbook("raghuvamsha_sarga_03.xlsx")
        ws = wb.active
        ws.cell(row=row_index+1, column=4).value = label_value
        wb.save("raghuvamsha_sarga_03.xlsx")

        # Update the dataframe with the new value
        df.loc[row_index, "Label"] = label_value

    # Render the template with the dataframe data
    return render_template("index.html", data=df[display_columns].values.tolist())

if __name__ == '__main__':
    app.run(debug=True)
